import os
import json
import uuid
import calendar
from datetime import datetime
from openpyxl import load_workbook
import logging


class ExcelProcessor:
    SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
    LOG_DIR = os.path.join(SCRIPT_DIR, "logs")
    LOG_FILE = os.path.join(LOG_DIR, "conversion_history.log")
    HISTORY_FILE = os.path.join(LOG_DIR, ".conversion_registry")

    def __init__(self, input_file, output_directory, force=False):
        self.input_file = os.path.abspath(input_file)
        self.output_directory = os.path.abspath(output_directory)
        self.force = force
        self.ensure_log_directory()
        self.setup_logging()

    def ensure_log_directory(self):
        os.makedirs(self.LOG_DIR, exist_ok=True)
        if os.name == "posix":
            try:
                import subprocess
                subprocess.run(["chattr", "+i", self.LOG_DIR], check=False)
            except:
                pass

    def setup_logging(self):
        formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
        self.logger = logging.getLogger("ExcelProcessor")
        self.logger.setLevel(logging.INFO)

        file_handler = logging.FileHandler(self.LOG_FILE)
        file_handler.setFormatter(formatter)
        self.logger.addHandler(file_handler)

        console_handler = logging.StreamHandler()
        console_handler.setFormatter(formatter)
        self.logger.addHandler(console_handler)

    def format_iso_date(self, excel_date):
        try:
            if isinstance(excel_date, str) and len(excel_date) == 7:
                year, month = excel_date.split("-")
                last_day = calendar.monthrange(int(year), int(month))[1]
                excel_date = f"{year}-{month}-{last_day}"
            return datetime.strptime(excel_date, "%Y-%m-%d").isoformat() + "-05:00"
        except (ValueError, TypeError):
            return None

    def get_file_tag(self, file_name):
        ext = os.path.splitext(file_name)[1]
        return ext[1:] if ext else ""

    def write_records_to_file(self, output_file_path, records, sheet_name):
        with open(output_file_path, "w") as json_file:
            for record in records:
                json_file.write(json.dumps(record, separators=(",", ":")) + "\n")
        self.logger.info(f"Created manifest file: {os.path.basename(output_file_path)} with {len(records)} records.")
        return len(records)

    def process_excel_sheet(self, sheet_name):
        self.logger.info(f"Processing sheet: {sheet_name}")
        wb = load_workbook(self.input_file, data_only=True)

        if sheet_name not in wb.sheetnames:
            self.logger.error(f"Sheet {sheet_name} not found in the workbook.")
            return

        sheet = wb[sheet_name]
        current_records = []
        records_per_file = 5000
        file_counter = 1
        publisher = sheet["B1"].value
        region = sheet["B4"].value
        record_class = sheet["C9"].value
        provenance = sheet["D1"].value
        security_classification = sheet["D3"].value
        creator = sheet["B2"].value
        contributor = sheet["D2"].value
        cost_center = str(sheet["B3"].value).zfill(12) if isinstance(sheet["B3"].value, (int, float)) else sheet["B3"].value.strip()

        for row in sheet.iter_rows(min_row=9, values_only=True):
            if row[0] is None:
                break
            file_name = row[0]
            file_path = row[1]
            folder_path = f"/dropzone/a360root/{publisher.lower()}/submission/{file_path}/"
            record_code = row[2]
            record_date = row[3]
            formatted_record_date = self.format_iso_date(record_date)
            relation_id = str(uuid.uuid4())
            submission_date = datetime.now().isoformat() + "-05:00"
            record_metadata = {
                "record_class": record_class,
                "publisher": publisher,
                "region": region,
                "recordDate": formatted_record_date,
                "provenance": provenance,
                "submission_date": submission_date,
                "security_classification": security_classification,
                "contributor": contributor,
                "creator": creator,
                "description": row[4],
                "title": file_name,
                "language": "eng",
                "cost_center": cost_center,
                "date_range": row[5],
                "major_description": row[6],
                "minor_description": row[7],
                "reference_1": row[8],
            }
            file_metadata = {
                "publisher": publisher,
                "source_folder_path": file_path,
                "source_file_name": file_name,
                "dz_folder_path": folder_path,
                "dz_file_name": file_name,
                "file_tag": self.get_file_tag(file_name),
            }
            current_records.append({"operation": "create_record", "relation_id": relation_id, "record_metadata": record_metadata})
            current_records.append({"operation": "upload_new_file", "relation_id": relation_id, "file_metadata": file_metadata})

            if len(current_records) >= records_per_file * 2:
                output_file_path = os.path.join(
                    self.output_directory,
                    f"{os.path.splitext(os.path.basename(self.input_file))[0]}_{sheet_name}_{file_counter}.a360"
                )
                self.write_records_to_file(output_file_path, current_records, sheet_name)
                current_records = []
                file_counter += 1

        if current_records:
            output_file_path = os.path.join(
                self.output_directory,
                f"{os.path.splitext(os.path.basename(self.input_file))[0]}_{sheet_name}_{file_counter}.a360"
            )
            self.write_records_to_file(output_file_path, current_records, sheet_name)

def main():
    import argparse

    parser = argparse.ArgumentParser(description="Process a specific sheet in an Excel file.")
    parser.add_argument("input_file", type=str, help="Path to the Excel file")
    parser.add_argument("output_directory", type=str, help="Directory to save the output files")
    parser.add_argument("sheet_name", type=str, help="Name of the sheet to process")
    parser.add_argument("--force", action="store_true", help="Force process even if already converted")

    args = parser.parse_args()

    os.makedirs(args.output_directory, exist_ok=True)
    processor = ExcelProcessor(args.input_file, args.output_directory, args.force)
    processor.process_excel_sheet(args.sheet_name)

if __name__ == "__main__":
    main()
