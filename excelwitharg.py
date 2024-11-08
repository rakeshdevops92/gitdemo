import os
import json
import uuid
import calendar
from datetime import datetime
from openpyxl import load_workbook
import argparse

def format_iso_date(excel_date):
    try:
        if isinstance(excel_date, str):
            if len(excel_date) == 7:
                year, month = excel_date.split('-')
                last_day = calendar.monthrange(int(year), int(month))[1]
                excel_date = f"{year}-{month}-{last_day}"
            return datetime.strptime(excel_date, "%Y-%m-%d").isoformat() + "-05:00"
    except (ValueError, TypeError):
        return None

def get_file_tag(file_name):
    ext = os.path.splitext(file_name)[1].lower()
    if not ext:
        ext = file_name.split(".")[-1]
    else:
        ext = ext[1:]  # Remove the leading dot
    return ext

def write_records_to_file(output_file_path, records):
    with open(output_file_path, 'w') as json_file:
        for record in records:
            json_file.write(json.dumps(record, separators=(',', ':')) + '\n')

def process_excel_file(file_path, output_directory):
    print(f"Processing file: {file_path}")
    wb = load_workbook(file_path, data_only=True)
    
    excel_base_name = os.path.splitext(os.path.basename(file_path))[0]
    file_count = 0
    records = []
    current_records = []
    records_per_file = 100
    
    for sheet_name in wb.sheetnames:
        if not sheet_name.startswith("Metadata"):
            continue
        print(f"Processing sheet: {sheet_name}")
        sheet = wb[sheet_name]

        publisher = sheet['B1'].value
        region = sheet['B4'].value
        record_class = sheet['C9'].value
        provenance = sheet['D1'].value
        security_classification = sheet['D4'].value
        security_classification_map = {
            "Confidential": "C",
            "Highly Confidential": "HC",
            "Internal": "I",
            "Public": "P"
        }
        security_classification = security_classification_map.get(security_classification, "I")

        creator = sheet['B2'].value
        contributor = sheet['D2'].value
        cost_center = str(sheet['B3'].value).zfill(12) if isinstance(sheet['B3'].value, (int, float)) else sheet['B3'].value.strip()

        for row in sheet.iter_rows(min_row=9, values_only=True):
            if row[0] is None:
                break

            file_name = row[0]
            file_path = row[1]
            folder_path = f"/dropzone/a360root/{publisher.lower()}/submission/{file_path}/"
            record_code = row[2]
            record_date = row[3]
            formatted_record_date = format_iso_date(record_date)
            relation_id = str(uuid.uuid4())
            submission_date = datetime.now().isoformat() + "-05:00"

            record_metadata = {
                "record_class": record_class,
                "publisher": publisher,
                "region": region,
                "recordDate": formatted_record_date,
                "provenance": provenance,
                "submission_date": submission_date,
                "security_classification": security_classification,
                "contributor": contributor,
                "creator": creator,
                "description": row[4],
                "title": file_name,
                "Language": "eng",
                "cost_center": cost_center,
                "date_range": row[5],
                "major_description": row[6],
                "minor_description": row[7],
                "reference_1": row[8]
            }

            file_metadata = {
                "publisher": publisher,
                "source_folder_path": file_path,
                "source_file_name": file_name,
                "dz_folder_path": folder_path,
                "dz_file_name": file_name,
                "file_tag": get_file_tag(file_name)
            }

            current_records.append({
                "operation": "create_record",
                "relation_id": relation_id,
                "record_metadata": record_metadata
            })
            current_records.append({
                "operation": "upload_new_file",
                "relation_id": relation_id,
                "file_metadata": file_metadata
            })

            if len(current_records) >= records_per_file * 2:
                output_file_path = os.path.join(
                    output_directory, f"{excel_base_name}_{sheet_name}_{file_count}.a360"
                )
                write_records_to_file(output_file_path, current_records)
                current_records = []
                file_count += 1

    if current_records:
        output_file_path = os.path.join(
            output_directory, f"{excel_base_name}_{sheet_name}_{file_count}.a360"
        )
        write_records_to_file(output_file_path, current_records)

    print(f"Files created for {os.path.basename(file_path)}: {file_count}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Convert Excel file to JSON format.")
    parser.add_argument("input_file", type=str, help="Path to the Excel file.")
    parser.add_argument("output_directory", type=str, help="Directory to save JSON files.")
    args = parser.parse_args()

    input_file = args.input_file
    output_directory = args.output_directory

    if not os.path.exists(output_directory):
        os.makedirs(output_directory)

    process_excel_file(input_file, output_directory)
