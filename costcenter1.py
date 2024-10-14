import json
import hashlib
from openpyxl import load_workbook
import pandas as pd

src_file = "/home/pj72963/metadata_excel/Content_Metadata_by_Cost_Center.xlsx"

def excel_to_json():
    wb = load_workbook(filename=src_file, data_only=True)
    ws = wb.active

    metadata = {}
    i = 1
    while True:
        key_cell_A = ws.cell(row=i, column=1).value
        value_cell_B = ws.cell(row=i, column=2).value
        key_cell_C = ws.cell(row=i, column=3).value
        value_cell_D = ws.cell(row=i, column=4).value

        if key_cell_A is None and key_cell_C is None:
            break

        if key_cell_A and value_cell_B:
            key = str(key_cell_A).strip()
            value = str(value_cell_B).strip()

            # If the key is "Cost Center", pad the value with zeros as in the Excel file
            if key.lower() == "cost center":
                value = value.zfill(10)
            metadata[key] = value

        if key_cell_C and value_cell_D:
            metadata[str(key_cell_C).strip()] = str(value_cell_D).strip()

        i += 1

    metadata = {k: v for k, v in metadata.items() if pd.notna(v)}

    relation_id = hashlib.md5(str(metadata).encode()).hexdigest()

    file_data_start_row = i + 3

    df_files = pd.read_excel(src_file, skiprows=file_data_start_row - 1, dtype=str)
    df_files.fillna("", inplace=True)
    df_files.dropna(axis=1, how='all', inplace=True)

    json_output = []

    json_output.append({
        "operation": "create_record",
        "relation_id": relation_id,
        "record_metadata": metadata
    })

    for _, row in df_files.iterrows():
        file_metadata = {col_name: row[col_name] for col_name in df_files.columns if row[col_name] != ""}
        if file_metadata:
            json_output.append({
                "operation": "upload_new_file",
                "relation_id": relation_id,
                "file_metadata": file_metadata
            })

    # Use the padded Cost Center value for the output file name
    cost_center = metadata.get("Cost Center", "default") + ".a360"
    dest_file = f"/home/pj72963/metadata_json/{cost_center}.json"

    with open(dest_file, 'w') as json_file:
        json.dump(json_output, json_file, indent=4)

    print(f"JSON file created successfully at {dest_file}")

excel_to_json()
